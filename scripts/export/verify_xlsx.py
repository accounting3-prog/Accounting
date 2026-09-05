"""
Opens the exported workbooks with openpyxl and checks what is really in them.

This shares no code with the exporter. openpyxl is a separate implementation of
the xlsx format, so if it can open the file and finds the right values in the
right cells, the file is genuinely a valid workbook and not merely bytes our own
writer happens to produce and our own reader happens to accept.

    python scripts/export/verify_xlsx.py <dir-written-by-build_workbooks.mjs>
"""

import csv
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

# Supplier names include Arabic; a Windows console defaults to cp1252.
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

out = Path(sys.argv[1])
expected = json.loads((out / "expected.json").read_text(encoding="utf-8"))
cards = expected["cards"]
txns = expected["transactions"]

failures = []


def check(label, ok, detail=""):
    if not ok:
        failures.append(label)
    print(("PASS  " if ok else "FAIL  ") + label + (("  - " + detail) if detail else ""))


# ------------------------------------------------------------ by-card workbook

wb = load_workbook(out / "by_card.xlsx")
names = wb.sheetnames

check("openpyxl opens the by-card workbook", True)
check(
    "there is one tab per card plus a summary",
    len(names) == len(cards) + 1,
    str(len(names)) + " tabs: " + str(names),
)
check("the first tab is the summary", names[0] == "Summary", names[0])
check(
    "every tab name is within Excel's 31-character limit",
    all(len(n) <= 31 for n in names),
    str([n for n in names if len(n) > 31]),
)
BAD = set(':\\/?*[]')
check(
    "no tab name uses a character Excel forbids",
    not any(set(n) & BAD for n in names),
    str([n for n in names if set(n) & BAD]),
)
check("tab names are unique", len(set(names)) == len(names))


def header_row(ws):
    """Find the header by looking for the row whose first cell says 'Card'."""
    for r in range(1, 12):
        if ws.cell(row=r, column=1).value == "Card":
            return r
    raise AssertionError("no header row in " + ws.title)


def data_rows(ws):
    h = header_row(ws)
    cols = [c.value for c in ws[h]]
    rows = []
    for r in range(h + 1, ws.max_row + 1):
        values = [ws.cell(row=r, column=i + 1).value for i in range(len(cols))]
        if all(v in (None, "") for v in values):
            continue
        rows.append(dict(zip(cols, values)))
    return rows


# Each card's tab must hold that card's transactions - all of them, and nothing
# belonging to another card. That is the whole point of the feature. Tabs are
# written in card order after the summary, so position identifies them without
# having to reimplement the name-sanitising rules here.
by_card_seen = {}
for i, card in enumerate(cards):
    tab = names[1 + i]
    rows = data_rows(wb[tab])
    mine = [t for t in txns if t["cardId"] == card["id"]]
    check(
        "'" + tab + "' holds its " + str(len(mine)) + " transaction(s)",
        len(rows) == len(mine),
        "found " + str(len(rows)),
    )
    check(
        "'" + tab + "' contains no other card's rows",
        all(r["Card"] == card["name"] for r in rows),
        str(sorted({str(r["Card"]) for r in rows})),
    )
    by_card_seen[card["id"]] = rows

total_in_tabs = sum(len(v) for v in by_card_seen.values())
check(
    "every transaction appears exactly once across the card tabs",
    total_in_tabs == len(txns),
    str(total_in_tabs) + " vs " + str(len(txns)),
)

# A card with no activity still gets a tab, so nobody wonders if it was dropped.
empty_i = next(i for i, c in enumerate(cards) if c["count"] == 0)
check(
    "a card with no transactions still has its own, empty tab",
    len(by_card_seen[cards[empty_i]["id"]]) == 0,
    names[1 + empty_i],
)

# Dates and amounts must survive as data, not as text needing to be re-parsed.
for t in txns:
    match = [
        r
        for r in by_card_seen[t["cardId"]]
        if r["Date"] == t["date"] and r["AED settlement"] == t["amount"]
    ]
    check(
        t["id"] + ": " + t["date"] + " " + format(t["amount"], ".3f") + " is present and unaltered",
        len(match) == 1,
        str(len(match)) + " matches",
    )
    if match:
        check(
            t["id"] + ": the AED amount is a number, not text",
            isinstance(match[0]["AED settlement"], (int, float)),
            type(match[0]["AED settlement"]).__name__,
        )

# Dates are written as ISO text so sorting is unambiguous in every locale.
all_dates = [r["Date"] for rows in by_card_seen.values() for r in rows]
check(
    "every date reads back as YYYY-MM-DD",
    all(
        isinstance(d, str) and len(d) == 10 and d[4] == "-" and d[7] == "-"
        for d in all_dates
    ),
    str(all_dates),
)
check(
    "no date was shifted or transposed",
    sorted(all_dates) == sorted(t["date"] for t in txns),
)

# Text that would break naive XML: an ampersand, angle brackets, quotes, a
# newline, and a non-Latin script.
suppliers = {r["Supplier"] for rows in by_card_seen.values() for r in rows}
check("an ampersand and angle brackets survive", "Smith & Sons <Travel>" in suppliers, str(suppliers))
check("double quotes survive", 'Refund "partial"' in suppliers)
check("a non-Latin script survives", "مورد عربي" in suppliers)
check(
    "an embedded newline survives",
    any("\n" in s for s in suppliers if isinstance(s, str)),
)

# ---------------------------------------------------------------- summary tab

summary = wb["Summary"]
shead = None
for r in range(1, 12):
    if summary.cell(row=r, column=1).value == "Card":
        shead = r
        break
check("the summary tab has a header row", shead is not None)
if shead:
    srows = [
        [summary.cell(row=r, column=c).value for c in range(1, 11)]
        for r in range(shead + 1, shead + 1 + len(cards))
    ]
    check(
        "the summary lists every card, under its real name",
        [row[0] for row in srows] == [c["name"] for c in cards],
        str([row[0] for row in srows]),
    )
    check(
        "the summary's transaction counts match the tabs",
        [row[1] for row in srows] == [c["count"] for c in cards],
        str([row[1] for row in srows]),
    )
    check(
        "the summary's ledger balances match the cards",
        [row[4] for row in srows] == [c["ledgerBalance"] for c in cards],
        str([row[4] for row in srows]),
    )
    check(
        "opening dates are shown where they exist",
        srows[0][8] == cards[0]["openingDate"],
        str(srows[0][8]),
    )
    check(
        "a card with no opening date shows blank, not a wrong one",
        srows[empty_i][8] in (None, ""),
        str(srows[empty_i][8]),
    )
    texts = [
        str(summary.cell(row=r, column=c).value or "")
        for r in range(1, summary.max_row + 1)
        for c in range(1, 12)
    ]
    check("the summary says which currency the balances are in", any("AED" in t for t in texts))

# -------------------------------------------------------- single-sheet export

wb2 = load_workbook(out / "single.xlsx")
check(
    "the single-sheet export still opens and is unchanged",
    wb2.sheetnames == ["Transactions"],
    str(wb2.sheetnames),
)
rows2 = data_rows(wb2["Transactions"])
check("the single sheet holds every transaction", len(rows2) == len(txns), str(len(rows2)))

# ------------------------------------------------------------------- the CSV

text = (out / "single.csv").read_text(encoding="utf-8-sig")
reader = list(csv.reader(text.splitlines()))
body = [r for r in reader if len(r) > 5 and r[0] != "Card"]
check("the CSV holds every transaction", len(body) == len(txns), str(len(body)))

print()
if failures:
    print(str(len(failures)) + " FAILED")
    for f in failures:
        print("  - " + f)
    sys.exit(1)
print("All checks passed.")

"""
Independent re-audit of the source workbook, from first principles.

This script deliberately shares NO code with extract.py. It rediscovers every
structural fact by a different method, so that agreement between the two is
evidence rather than tautology:

    fact                extract.py                  audit.py
    ----------------------------------------------------------------------
    header row          hardcoded per sheet         found by scoring rows
                                                    against header keywords
    balance column      hardcoded per sheet         found as the column whose
                                                    formulas reference the cell
                                                    above in their own column
    money columns       hardcoded per sheet         read out of the balance
                                                    formula's references
    direction           parses the formula's        arithmetic on Excel's
                        +/- signs                   cached balance deltas
    balance chain       replays parsed amounts      replays cell values read
                                                    straight from the sheet

Where the two disagree about anything, that disagreement is the finding.

The workbook is opened read-only and never saved. Nothing here writes to a
database. Nothing is corrected, dropped, merged or guessed: every exception is
reported with its sheet, row and the raw value that caused it.
"""

import argparse
import datetime
import json
import os
import re
import sys
import warnings
from collections import Counter, defaultdict

import openpyxl
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

WORKBOOK = r"C:/Users/LE.Andrew/Downloads/2026 Cards Monitoring.xlsx"

HEADER_KEYWORDS = {
    "date", "details", "supplier", "description", "currency", "debit", "credit",
    "balance", "conversion", "crm", "req", "request", "lpo", "invoice",
    "payment", "account", "client", "commissionable", "sales", "amount",
}

ISO_4217 = {
    "AED", "USD", "EUR", "GBP", "SAR", "JPY", "EGP", "CHF", "TRY", "OMR",
    "QAR", "BHD", "KRW", "MYR", "SGD", "KWD", "VND", "HKD", "MOP", "JOD",
    "INR", "ZAR", "MAD", "SEK", "CZK", "MUR", "CAD", "NZD",
}

TOLERANCE = 0.005


class Findings:
    """Every exception, kept with enough context to act on."""

    def __init__(self):
        self.items = []

    def add(self, sheet, row, kind, detail, raw=None):
        self.items.append({"sheet": sheet, "row": row, "kind": kind,
                           "detail": detail, "raw": raw})

    def by_kind(self):
        out = defaultdict(list)
        for f in self.items:
            out[f["kind"]].append(f)
        return out

    def of(self, sheet, kind=None):
        return [f for f in self.items
                if f["sheet"] == sheet and (kind is None or f["kind"] == kind)]


# ---------------------------------------------------------------------------
# Structure discovery — none of this is told to us
# ---------------------------------------------------------------------------

def find_header_row(ws):
    """The row scoring highest on header-like words, searching the top 10."""
    best, best_score = None, 0
    for r in range(1, min(11, ws.max_row + 1)):
        score = 0
        for c in range(1, min(ws.max_column, 30) + 1):
            v = ws.cell(r, c).value
            if not isinstance(v, str):
                continue
            words = re.findall(r"[a-z]+", v.lower())
            if any(w in HEADER_KEYWORDS for w in words):
                score += 1
        if score > best_score:
            best, best_score = r, score
    return best, best_score


def find_balance_columns(ws, header_row, findings, sheet):
    """Find every running-total column, then choose the account's own.

    Two traps in this workbook, both of which a naive "most formulas wins"
    rule gets wrong:

      * MASTERCARD 5135 carries TWO chains — column G under the BALANCE header
        (the account), and an unlabelled column R with MORE formulas that
        tracks something else against a 250,000 figure. Formula count picks the
        wrong one, so a column headed 'balance' wins over one that is not.
      * AMEX 2044's only chain formula hardcodes its opening balance
        (`=13709.16-D6+E6`) rather than referencing the cell above, so a
        strict self-reference test finds no balance column at all.
    """
    self_ref, seeded = Counter(), Counter()
    for c in range(1, ws.max_column + 1):
        letter = get_column_letter(c)
        for r in range(header_row + 1, ws.max_row + 1):
            v = ws.cell(r, c).value
            if not (isinstance(v, str) and v.startswith("=")):
                continue
            if re.search(rf"\b{letter}{r - 1}\b", v):
                self_ref[letter] += 1
            elif re.match(r"=\s*[\d.]+\s*[-+]", v) and re.search(r"[A-Z]+\d+", v):
                # Opening balance written into the formula as a literal.
                seeded[letter] += 1

    candidates = Counter(self_ref)
    for letter, n in seeded.items():
        if letter not in candidates:
            candidates[letter] = n
            findings.add(sheet, header_row, "balance_formula_hardcodes_opening",
                         f"column {letter} carries the running balance but writes "
                         f"the opening figure into the formula as a literal "
                         f"instead of referencing the cell above",
                         ws[f"{letter}{header_row + 2}"].value)
    if not candidates:
        return None, 0, []

    def headed_balance(letter):
        return "balance" in str(ws[f"{letter}{header_row}"].value or "").lower()

    ranked = sorted(candidates.items(),
                    key=lambda kv: (headed_balance(kv[0]), kv[1]), reverse=True)
    chosen, count = ranked[0]

    others = [l for l, n in ranked[1:] if n >= 3]
    for letter in others:
        findings.add(sheet, header_row, "secondary_balance_chain",
                     f"column {letter} runs a second balance chain "
                     f"({candidates[letter]} formulas, header "
                     f"{ws[f'{letter}{header_row}'].value!r}) alongside the "
                     f"account balance in column {chosen}; not imported", None)
    return chosen, count, others


def money_columns_from_formula(ws, header_row, bal_col):
    """Read the two money columns out of the balance formula itself."""
    refs = Counter()
    for r in range(header_row + 1, ws.max_row + 1):
        v = ws.cell(r, openpyxl.utils.column_index_from_string(bal_col)).value
        if not (isinstance(v, str) and v.startswith("=")):
            continue
        for col, row in re.findall(r"([A-Z]+)(\d+)", v):
            if col != bal_col and int(row) == r:
                refs[col] += 1
    return [c for c, _ in refs.most_common(2)]


def direction_from_cached_deltas(ws_vals, header_row, bal_col, money_cols):
    """Which column decreases the balance, decided purely by arithmetic on the
    values Excel cached. No formula text is read."""
    votes = Counter()
    prev = None
    for r in range(header_row + 1, ws_vals.max_row + 1):
        cur = ws_vals[f"{bal_col}{r}"].value
        if not isinstance(cur, (int, float)):
            continue
        if prev is not None:
            delta = float(cur) - float(prev)
            for col in money_cols:
                v = ws_vals[f"{col}{r}"].value
                if isinstance(v, (int, float)) and abs(v) > 0:
                    if abs(delta + float(v)) < TOLERANCE:
                        votes[(col, "decreases")] += 1
                    elif abs(delta - float(v)) < TOLERANCE:
                        votes[(col, "increases")] += 1
        prev = cur
    dec = {c for (c, k), n in votes.items() if k == "decreases" and n > 0}
    inc = {c for (c, k), n in votes.items() if k == "increases" and n > 0}
    return votes, dec, inc


def balance_formula_pattern(ws, header_row, bal_col):
    shapes = Counter()
    for r in range(header_row + 1, ws.max_row + 1):
        v = ws.cell(r, openpyxl.utils.column_index_from_string(bal_col)).value
        if isinstance(v, str) and v.startswith("="):
            shapes[re.sub(r"\d+", "#", v)] += 1
    return shapes


# ---------------------------------------------------------------------------
# Value auditing
# ---------------------------------------------------------------------------

def audit_dates(ws, header_row, date_col, sheet, findings):
    """Classify every date cell without repairing anything."""
    stats = Counter()
    text_first_parts = []
    ambiguous_rows, confirmed = [], []

    for r in range(header_row + 1, ws.max_row + 1):
        v = ws[f"{date_col}{r}"].value
        if v is None or (isinstance(v, str) and not v.strip()):
            continue
        if isinstance(v, datetime.datetime):
            d = v.date()
            stats["datetime"] += 1
            if d.day > 12:
                stats["datetime_unambiguous"] += 1
                confirmed.append((r, d))
            else:
                stats["datetime_ambiguous"] += 1
                ambiguous_rows.append((r, d))
        elif isinstance(v, str):
            s = v.strip()
            m = re.fullmatch(r"(\d{1,2})[/-](\d{1,2})[/-](\d{4})", s)
            if m:
                stats["text"] += 1
                text_first_parts.append(int(m.group(1)))
                try:
                    confirmed.append(
                        (r, datetime.date(int(m.group(3)), int(m.group(2)),
                                          int(m.group(1)))))
                except ValueError:
                    findings.add(sheet, r, "date_invalid",
                                 "text date is not a valid day/month/year", s)
            else:
                stats["text_unparseable"] += 1
                findings.add(sheet, r, "date_unparseable",
                             "date cell is text in no recognised format", s)
        else:
            stats["other_type"] += 1
            findings.add(sheet, r, "date_unexpected_type",
                         f"date cell holds a {type(v).__name__}", repr(v))

    # Decide, per sheet, whether the ambiguous cells read correctly as-is or
    # only make chronological sense transposed. Each ambiguous cell is bracketed
    # between the nearest confirmed dates either side.
    ordered = sorted(confirmed + [(r, d) for r, d in ambiguous_rows])
    conf_map = dict(confirmed)
    rows_sorted = [r for r, _ in ordered]
    asis = swap = neither = 0
    for r, d in ambiguous_rows:
        i = rows_sorted.index(r)
        prev = next((conf_map[rows_sorted[j]] for j in range(i - 1, -1, -1)
                     if rows_sorted[j] in conf_map), None)
        nxt = next((conf_map[rows_sorted[j]] for j in range(i + 1, len(rows_sorted))
                    if rows_sorted[j] in conf_map), None)
        try:
            sw = datetime.date(d.year, d.day, d.month)
        except ValueError:
            sw = None

        def fits(x):
            return x is not None and (prev is None or x >= prev) \
                and (nxt is None or x <= nxt)

        f1, f2 = fits(d), fits(sw)
        if f1 and not f2:
            asis += 1
        elif f2 and not f1:
            swap += 1
        else:
            neither += 1

    return {
        "counts": dict(stats),
        "text_first_part_all_over_12": all(p > 12 for p in text_first_parts)
                                       if text_first_parts else None,
        "text_date_count": len(text_first_parts),
        "ambiguous": len(ambiguous_rows),
        "fits_as_is": asis, "fits_swapped": swap, "fits_neither": neither,
        "verdict": ("transposed" if swap > 0 and asis == 0 else
                    "as_written" if asis > 0 and swap == 0 else
                    "none" if not ambiguous_rows else "mixed"),
    }


CURRENCY_RE = re.compile(r"'?([A-Za-z]{3})\s*([\d,]*\.?\d*)\.?'?")


def audit_currency_and_rates(ws, ws_vals, header_row, sheet, cols, findings):
    """Check every currency cell and every conversion formula."""
    curr_col = cols.get("currency")
    code_col, amt_col = cols.get("currency_code"), cols.get("currency_amount")
    conv_col = cols.get("conversion")

    codes = Counter()
    rate_rows = 0
    for r in range(header_row + 1, ws.max_row + 1):
        code = orig = None

        if code_col:  # split across two columns
            cv = ws[f"{code_col}{r}"].value
            av = ws[f"{amt_col}{r}"].value
            if cv is not None and str(cv).strip():
                code = str(cv).strip().upper()
                if code not in ISO_4217:
                    findings.add(sheet, r, "currency_unrecognised",
                                 "currency code is not a known ISO 4217 code", str(cv))
                    code = None
                orig = av if isinstance(av, (int, float)) else None
        elif curr_col:
            v = ws[f"{curr_col}{r}"].value
            if v is not None and str(v).strip():
                s = str(v).strip()
                m = CURRENCY_RE.fullmatch(s)
                if not m:
                    findings.add(sheet, r, "currency_unparseable",
                                 "currency cell has no readable 3-letter code", s)
                else:
                    code = m.group(1).upper()
                    if code not in ISO_4217:
                        findings.add(sheet, r, "currency_unrecognised",
                                     "code is not a known ISO 4217 code", s)
                        code = None
                    if m.group(2):
                        try:
                            orig = float(m.group(2).replace(",", ""))
                        except ValueError:
                            findings.add(sheet, r, "currency_amount_unparseable",
                                         "amount beside the code is not a number", s)
        if code:
            codes[code] += 1

        if not conv_col:
            continue
        f = ws[f"{conv_col}{r}"].value
        if not (isinstance(f, str) and f.startswith("=")):
            continue
        rate_rows += 1
        cached = ws_vals[f"{conv_col}{r}"].value

        simple = re.fullmatch(r"=\s*([A-Z]+)(\d+)\s*/\s*([\d,.]+)\s*", f)
        if not simple:
            findings.add(sheet, r, "rate_formula_hardcoded",
                         "conversion hardcodes both sides, so the original "
                         "amount cannot be recovered from it", f)
            continue

        num_col, num_row, den = simple.group(1), int(simple.group(2)), simple.group(3)
        if num_row != r:
            findings.add(sheet, r, "rate_row_mismatch",
                         f"conversion divides a value from row {num_row}", f)
        try:
            den_f = float(den.replace(",", ""))
        except ValueError:
            findings.add(sheet, r, "rate_denominator_unparseable",
                         "denominator is not a number", f)
            continue

        if orig is not None and abs(den_f - orig) > 1e-9:
            findings.add(sheet, r, "rate_denominator_mismatch",
                         f"conversion divides by {den} but the currency cell "
                         f"states {orig}", f)
        if orig is None and code is None:
            findings.add(sheet, r, "rate_without_currency",
                         "a conversion rate exists but the currency cell could "
                         "not be read", f)

        # Independently recompute the rate and compare with Excel's cached one.
        num_v = ws_vals[f"{num_col}{num_row}"].value
        if isinstance(num_v, (int, float)) and den_f and isinstance(cached, (int, float)):
            if abs(float(num_v) / den_f - float(cached)) > 1e-6:
                findings.add(sheet, r, "rate_cached_value_wrong",
                             f"cached rate {cached} does not equal "
                             f"{num_v}/{den_f}", f)
    return dict(codes), rate_rows


# ---------------------------------------------------------------------------
# The per-sheet audit
# ---------------------------------------------------------------------------

def audit_sheet(wb, wbv, name, findings):
    ws, wsv = wb[name], wbv[name]

    header_row, header_score = find_header_row(ws)
    bal_col, bal_count, secondary = find_balance_columns(
        ws, header_row, findings, name)
    if not bal_col:
        findings.add(name, 0, "no_balance_column",
                     "no column looks like a running balance", None)
        return None

    money_cols = money_columns_from_formula(ws, header_row, bal_col)
    votes, dec_set, inc_set = direction_from_cached_deltas(
        wsv, header_row, bal_col, money_cols)
    shapes = balance_formula_pattern(ws, header_row, bal_col)

    if len(dec_set) > 1:
        findings.add(name, header_row, "ambiguous_direction",
                     f"more than one column appears to decrease the balance: "
                     f"{sorted(dec_set)}", None)
    if len(shapes) > 1:
        findings.add(name, header_row, "multiple_balance_patterns",
                     f"balance column uses {len(shapes)} different formula "
                     f"shapes: {list(shapes)}", None)

    dec_col = sorted(dec_set)[0] if dec_set else None
    inc_col = next((c for c in money_cols if c != dec_col), None)

    # Header labels, purely for the record — never used to decide direction.
    dec_header = ws[f"{dec_col}{header_row}"].value if dec_col else None
    inc_header = ws[f"{inc_col}{header_row}"].value if inc_col else None
    if dec_header and "credit" in str(dec_header).lower():
        findings.add(name, header_row, "misleading_header",
                     f"the column that DECREASES the balance is labelled "
                     f"{dec_header!r}", None)

    # Opening balance and every later hardcoded value (a manual overwrite).
    seeds = [(r, float(ws[f"{bal_col}{r}"].value))
             for r in range(header_row + 1, ws.max_row + 1)
             if isinstance(ws[f"{bal_col}{r}"].value, (int, float))]
    opening_row, opening_balance = seeds[0] if seeds else (None, None)
    for r, val in seeds[1:]:
        findings.add(name, r, "manual_balance_overwrite",
                     f"balance typed as {val:,.2f} instead of carried by "
                     f"formula", str(val))

    # Rows carrying money, and rows carrying money with no balance formula.
    txn_rows, missing_formula = [], []
    for r in range(header_row + 1, ws.max_row + 1):
        if r == opening_row:
            continue
        has_money = any(isinstance(ws[f"{c}{r}"].value, (int, float))
                        for c in money_cols)
        if not has_money:
            continue
        txn_rows.append(r)
        bv = ws[f"{bal_col}{r}"].value
        if not (isinstance(bv, str) and bv.startswith("=")) \
                and not isinstance(bv, (int, float)):
            missing_formula.append(r)
            findings.add(name, r, "missing_balance_formula",
                         "row carries an amount but no balance formula, so the "
                         "workbook's own balance excludes it",
                         str(ws[f"B{r}"].value))

    # Independent chain recomputation, straight from cell values.
    running, checked, chain_breaks = opening_balance, 0, []
    for r in txn_rows:
        if r in missing_formula:
            continue
        dv = ws[f"{dec_col}{r}"].value if dec_col else None
        iv = ws[f"{inc_col}{r}"].value if inc_col else None
        step = 0.0
        if isinstance(dv, (int, float)):
            step -= float(dv)
        if isinstance(iv, (int, float)):
            step += float(iv)
        running += step
        cached = wsv[f"{bal_col}{r}"].value
        if isinstance(ws[f"{bal_col}{r}"].value, (int, float)):
            running = float(ws[f"{bal_col}{r}"].value)   # manual overwrite
            continue
        if isinstance(cached, (int, float)):
            checked += 1
            if abs(running - float(cached)) > TOLERANCE:
                chain_breaks.append({"row": r, "computed": round(running, 2),
                                     "cached": round(float(cached), 2)})
                findings.add(name, r, "balance_chain_mismatch",
                             f"recomputed {running:,.2f} but the sheet caches "
                             f"{float(cached):,.2f}", None)
                running = float(cached)

    # The workbook's own closing balance: the last row that has both an amount
    # and a balance figure.
    last_balance = None
    for r in txn_rows:
        bv = wsv[f"{bal_col}{r}"].value
        if isinstance(bv, (int, float)):
            last_balance = float(bv)

    # Dates
    date_col = get_column_letter(1)
    dates = audit_dates(ws, header_row, date_col, name, findings)

    # Currency and rates
    cols = {"conversion": None}
    if name == "RAK 9825 (6071)":
        cols.update({"currency_code": "C", "currency_amount": "D"})
    else:
        cols["currency"] = "C"
        conv_scores = Counter()
        for c in range(1, ws.max_column + 1):
            letter = get_column_letter(c)
            if letter == bal_col:
                continue
            for r in range(header_row + 1, ws.max_row + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str) and re.fullmatch(
                        r"=\s*[A-Z]+\d+\s*/\s*[\d,.]+\s*", v):
                    conv_scores[letter] += 1
        cols["conversion"] = conv_scores.most_common(1)[0][0] if conv_scores else None
    codes, rate_rows = audit_currency_and_rates(
        ws, wsv, header_row, name, cols, findings)

    # Text fields and duplicate candidates
    supplier_blank = 0
    sig = Counter()
    for r in txn_rows:
        b = ws[f"B{r}"].value
        if not (b and str(b).strip()):
            supplier_blank += 1
            findings.add(name, r, "supplier_missing",
                         "row has an amount but no supplier or description", None)
        dv = ws[f"{dec_col}{r}"].value if dec_col else None
        iv = ws[f"{inc_col}{r}"].value if inc_col else None
        amt = -float(dv) if isinstance(dv, (int, float)) else (
            float(iv) if isinstance(iv, (int, float)) else None)
        dcell = ws[f"{date_col}{r}"].value
        sig[(str(dcell), str(b).strip().upper() if b else "", amt)] += 1
    dup_groups = {k: n for k, n in sig.items() if n > 1}
    for (d, s, a), n in sorted(dup_groups.items(), key=lambda kv: -kv[1])[:200]:
        findings.add(name, 0, "duplicate_candidate",
                     f"{n} rows share date/supplier/amount: {s[:40]!r} "
                     f"{a} on {d[:10]}", None)

    return {
        "sheet": name,
        "header_row": header_row,
        "header_confidence": header_score,
        "balance_column": bal_col,
        "balance_formula_count": bal_count,
        "secondary_balance_chains": secondary,
        "balance_formula_patterns": dict(shapes),
        "money_columns": money_cols,
        "decreasing_column": dec_col,
        "decreasing_header": dec_header,
        "increasing_column": inc_col,
        "increasing_header": inc_header,
        "header_is_misleading": bool(dec_header and "credit" in str(dec_header).lower()),
        "opening_balance": opening_balance,
        "opening_row": opening_row,
        "manual_overwrites": [{"row": r, "value": v} for r, v in seeds[1:]],
        "transaction_rows": len(txn_rows),
        "rows_missing_balance_formula": missing_formula,
        "chain_rows_checked": checked,
        "chain_breaks": chain_breaks,
        "last_formula_chain_balance": last_balance,
        "recomputed_final": round(running, 2) if running is not None else None,
        "dates": dates,
        "currency_column": cols.get("currency") or cols.get("currency_code"),
        "conversion_column": cols.get("conversion"),
        "currency_codes": codes,
        "rate_rows": rate_rows,
        "duplicate_candidate_groups": len(dup_groups),
        "duplicate_candidate_rows": sum(dup_groups.values()),
        "supplier_missing": supplier_blank,
    }


# ---------------------------------------------------------------------------
# Cross-check against extract.py's output
# ---------------------------------------------------------------------------

def cross_check(sheets, normalised_path):
    if not os.path.exists(normalised_path):
        return None
    with open(normalised_path, encoding="utf-8") as fh:
        cards = {c["card"]: c for c in json.load(fh)}
    rows = []
    for s in sheets:
        c = cards.get(s["sheet"])
        if not c:
            rows.append({"sheet": s["sheet"], "issue": "missing from extraction"})
            continue
        src = [t for t in c["transactions"]
               if t["entry_type"] == "source_transaction"]
        rows.append({
            "sheet": s["sheet"],
            "header_row": (s["header_row"], c["header_row"],
                           s["header_row"] == c["header_row"]),
            "decreasing_column": (s["decreasing_column"], c["decreasing_column"],
                                  s["decreasing_column"] == c["decreasing_column"]),
            "increasing_column": (s["increasing_column"], c["increasing_column"],
                                  s["increasing_column"] == c["increasing_column"]),
            "opening_balance": (s["opening_balance"], c["opening_balance"],
                                abs((s["opening_balance"] or 0)
                                    - (c["opening_balance"] or 0)) < TOLERANCE),
            "transaction_count": (s["transaction_rows"], len(src),
                                  s["transaction_rows"] == len(src)),
            "source_balance": (s["last_formula_chain_balance"],
                               c["summary"]["source_balance"],
                               abs((s["last_formula_chain_balance"] or 0)
                                   - c["summary"]["source_balance"]) < TOLERANCE),
        })
    return rows


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--workbook", default=WORKBOOK)
    ap.add_argument("--normalised", default="scripts/out/normalised.json")
    ap.add_argument("--out", help="write the full audit JSON here")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.workbook, data_only=False)
    wbv = openpyxl.load_workbook(args.workbook, data_only=True)
    findings = Findings()

    print("=" * 92)
    print("INDEPENDENT RE-AUDIT — structure rediscovered from the workbook itself")
    print(f"workbook: {args.workbook}")
    print(f"sheets:   {len(wb.sheetnames)}")
    print("=" * 92)

    sheets = []
    for name in wb.sheetnames:
        s = audit_sheet(wb, wbv, name, findings)
        if s:
            sheets.append(s)

    print("\n1. STRUCTURE  (header row, balance column, direction)")
    print("-" * 92)
    for s in sheets:
        print(f"\n  {s['sheet']}")
        print(f"    header row {s['header_row']} (matched {s['header_confidence']} "
              f"header words)   balance column {s['balance_column']} "
              f"({s['balance_formula_count']} formulas)")
        pat = list(s["balance_formula_patterns"])[0] if s["balance_formula_patterns"] else "-"
        print(f"    formula pattern      {pat}"
              f"{'  [+%d more]' % (len(s['balance_formula_patterns']) - 1) if len(s['balance_formula_patterns']) > 1 else ''}")
        flag = "   << labelled CREDIT" if s["header_is_misleading"] else ""
        print(f"    DECREASES balance    col {s['decreasing_column']} "
              f"({s['decreasing_header']!r}){flag}")
        print(f"    INCREASES balance    col {s['increasing_column']} "
              f"({s['increasing_header']!r})")

    print("\n\n2. BALANCES  (chain recomputed independently from cell values)")
    print("-" * 92)
    print(f"  {'SHEET':<32}{'OPENING':>13}{'TXNS':>7}{'CHECKED':>9}"
          f"{'BREAKS':>8}{'CLOSING':>14}")
    for s in sheets:
        print(f"  {s['sheet']:<32}{s['opening_balance']:>13,.2f}"
              f"{s['transaction_rows']:>7}{s['chain_rows_checked']:>9}"
              f"{len(s['chain_breaks']):>8}"
              f"{(s['last_formula_chain_balance'] or 0):>14,.2f}")
    tot_checked = sum(s["chain_rows_checked"] for s in sheets)
    tot_breaks = sum(len(s["chain_breaks"]) for s in sheets)
    tot_txn = sum(s["transaction_rows"] for s in sheets)
    print(f"  {'TOTAL':<32}{'':>13}{tot_txn:>7}{tot_checked:>9}{tot_breaks:>8}")

    print("\n\n3. OPENING DATE  (defined as the earliest transaction on the card)")
    print("-" * 92)
    for s in sheets:
        d = s["dates"]
        print(f"  {s['sheet']:<32} dates: {d['counts']}")
        print(f"  {'':<32} ambiguous {d['ambiguous']}  "
              f"as-written {d['fits_as_is']}  transposed {d['fits_swapped']}  "
              f"neither {d['fits_neither']}   VERDICT: {d['verdict']}")

    print("\n\n4. CURRENCIES AND RATES")
    print("-" * 92)
    for s in sheets:
        top = dict(sorted(s["currency_codes"].items(),
                          key=lambda kv: -kv[1])[:6])
        print(f"  {s['sheet']:<32} col {s['currency_column']}  "
              f"conv col {s['conversion_column']}  rate rows {s['rate_rows']}")
        print(f"  {'':<32} {top}")

    print("\n\n5. FINDINGS  (nothing corrected, dropped, merged or guessed)")
    print("-" * 92)
    kinds = findings.by_kind()
    if not kinds:
        print("  none")
    for kind, items in sorted(kinds.items(), key=lambda kv: -len(kv[1])):
        print(f"\n  {kind}  ({len(items)})")
        for f in items[:6]:
            loc = f"row {f['row']}" if f["row"] else "sheet-level"
            raw = f"  raw={f['raw']!r}" if f["raw"] else ""
            print(f"      {f['sheet']} {loc}: {f['detail']}{raw}")
        if len(items) > 6:
            print(f"      ... and {len(items) - 6} more")

    print("\n\n6. CROSS-CHECK AGAINST extract.py")
    print("-" * 92)
    xc = cross_check(sheets, args.normalised)
    if xc is None:
        print("  (no extraction output found to compare against)")
    else:
        disagreements = 0
        for row in xc:
            bad = [k for k, v in row.items()
                   if isinstance(v, tuple) and not v[2]]
            if bad:
                disagreements += len(bad)
                print(f"  {row['sheet']}: DISAGREES on {bad}")
                for k in bad:
                    print(f"      {k}: audit={row[k][0]!r} extract={row[k][1]!r}")
            else:
                print(f"  {row['sheet']}: agrees on every checked fact")
        print(f"\n  total disagreements: {disagreements}")

    if args.out:
        os.makedirs(os.path.dirname(args.out), exist_ok=True)
        with open(args.out, "w", encoding="utf-8") as fh:
            json.dump({"sheets": sheets, "findings": findings.items},
                      fh, indent=2, ensure_ascii=False, default=str)
        print(f"\naudit written -> {args.out}")

    blocking = [f for f in findings.items
                if f["kind"] in ("balance_chain_mismatch", "no_balance_column",
                                 "ambiguous_direction", "rate_cached_value_wrong")]
    print(f"\nBLOCKING findings (must be resolved before import): {len(blocking)}")
    for f in blocking:
        print(f"  {f['sheet']} row {f['row']}: {f['kind']} — {f['detail']}")
    return 1 if blocking else 0


if __name__ == "__main__":
    sys.exit(main())

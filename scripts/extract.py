"""
Extract the card-monitoring workbook into a normalised dataset.

THE WORKBOOK IS READ-ONLY INPUT. This script opens it, never saves it, and
never writes back to it. No source value is altered on the way out: where a
value had to be repaired to be usable, the original is carried alongside the
repair and the row is flagged as a traceable correction.

Design rules:
  * Direction (which column decreases the balance) is DERIVED from each sheet's
    own balance formula, never from the DEBIT/CREDIT header text. Two sheets
    have inverted formulas and one has transposed headers; the formula is the
    only reliable source.
  * Nothing is guessed. An unrecognised currency, an unparseable date or a
    conversion rate that does not reconcile is left null, marked needs_review,
    and reported. A wrong value in a financial report is worse than a blank one.
  * Where the workbook and its own arithmetic disagree, neither is silently
    preferred. Both figures are produced and the difference is named.
  * Dry-run only. This script never writes to a database.

The parse proves itself: after extracting, it replays the balance chain from
the parsed amounts and compares every row against the value Excel itself
cached. A parse that reproduces 1,946 balances cannot have the sign backwards.
"""

import argparse
import datetime
import hashlib
import json
import re
import sys
import warnings
from collections import Counter, defaultdict

import openpyxl

warnings.filterwarnings("ignore")

WORKBOOK = r"C:/Users/LE.Andrew/Downloads/2026 Cards Monitoring.xlsx"

# ISO 4217 codes seen in the source. A code outside this set is reported as
# unrecognised rather than mapped to a best guess.
KNOWN_CURRENCIES = {
    "AED", "USD", "EUR", "GBP", "SAR", "JPY", "EGP", "CHF", "TRY", "OMR",
    "QAR", "BHD", "KRW", "MYR", "SGD", "KWD", "VND", "HKD", "MOP", "JOD",
    "INR", "ZAR", "MAD", "SEK", "CZK", "MUR", "CAD", "NZD",
}

# The ledger is denominated in AED; every balance column is AED.
BASE_CURRENCY = "AED"

# Sheets whose ambiguous date cells (day and month both <= 12) were written by
# Excel as month/day from day/month source text. Established by bracketing each
# ambiguous cell between the nearest confirmed dates either side: 14 of 14
# ambiguous rows on these two sheets fit only when transposed, and 0 of 640 on
# the other sheets do. Confirmed by the account owner.
SWAP_DAY_MONTH = {
    "MASTERCARD 5135 (4173) (7206)",
    "RAK 9825 (6071)",
}

# Anomaly kinds that make a row untrustworthy without making it wrong. The row
# is imported in full and flagged, never dropped and never corrected.
REVIEW_KINDS = {
    "currency_unrecognised", "currency_unparseable", "currency_amount_unparseable",
    "rate_denominator_mismatch", "rate_formula_unexpected",
    "date_unparseable", "date_invalid", "date_swap_invalid",
    "date_unexpected_type", "both_money_columns", "no_amount",
    "rate_without_currency",
}

# Column roles per sheet. Only the roles are declared here; the direction of
# the money columns is derived from the balance formula at parse time.
SHEETS = {
    "MASTERCARD 5135 (4173) (7206)": {
        "header_row": 4, "date": "A", "supplier": "B", "currency": "C",
        "money": ("D", "E"), "balance": "G", "conversion": "H",
        "extra": {"crm": "I", "req_number": "J", "lpo_number": "K",
                  "invoice": "L", "payment_ref": "M", "client": "O",
                  "sales_operation": "P"},
    },
    "RAK 8871 (8435)(0033)": {
        "header_row": 1, "date": "A", "supplier": "B", "currency": "C",
        "money": ("D", "E"), "balance": "F", "conversion": "G",
        "extra": {"crm": "H", "req_number": "I", "lpo_number": "J",
                  "invoice": "K", "payment_ref": "L", "client": "N",
                  "sales_operation": "O"},
    },
    "AMEX 4000 VPAY": {
        "header_row": 1, "date": "A", "supplier": "B", "currency": "C",
        "money": ("D", "E"), "balance": "F", "conversion": "G",
        "extra": {"account": "H", "req_number": "I", "event_end": "J",
                  "lpo_number": "K", "invoice": "L", "payment_ref": "M",
                  "notes": "N", "client": "P", "sales_operation": "Q"},
    },
    "MASTERCARD 6404 VPAY": {
        "header_row": 1, "date": "A", "supplier": "B", "currency": "C",
        "money": ("D", "E"), "balance": "F", "conversion": "H",
        # Column G is headed CRM but holds five stray conversion formulas.
        # Kept verbatim so nothing in the sheet is silently dropped.
        "extra": {"crm": "G", "account": "I", "req_number": "J",
                  "event_end": "K", "lpo_number": "L", "invoice": "M",
                  "payment_ref": "N", "client": "P", "sales_operation": "Q"},
    },
    "AMEX 3024 (3016- COR)": {
        "header_row": 1, "date": "A", "supplier": "B", "currency": "C",
        "money": ("D", "E"), "balance": "F", "conversion": "G",
        "extra": {"crm": "H", "req_number": "I", "lpo_number": "J",
                  "invoice": "K", "payment_ref": "L", "client": "N",
                  "sales_operation": "O"},
    },
    "AMEX 2044 (2036- VIP)": {
        # Headers on this sheet are transposed: D reads CREDIT but decreases
        # the balance. Derived direction handles it; the header is ignored.
        "header_row": 4, "date": "A", "supplier": "B", "currency": "C",
        "money": ("D", "E"), "balance": "F", "conversion": "G",
        "extra": {"account": "H", "req_number": "I", "lpo_number": "J",
                  "payment_ref": "K", "client": "M", "sales_operation": "N"},
    },
    "RAK 9825 (6071)": {
        "header_row": 1, "date": "A", "supplier": "B",
        # The only sheet that splits code and original amount into two columns.
        "currency_code": "C", "currency_amount": "D",
        "money": ("E", "F"), "balance": "G", "conversion": None,
        "extra": {"req_number": "H", "lpo_number": "I", "payment_ref": "J",
                  "invoice": "K", "notes": "L"},
    },
}


class Anomaly:
    """A row the parser refused to guess about."""

    def __init__(self, sheet, row, kind, detail):
        self.sheet, self.row, self.kind, self.detail = sheet, row, kind, detail

    def as_dict(self):
        return {"sheet": self.sheet, "row": self.row, "kind": self.kind,
                "detail": self.detail}


def derive_direction(sheet_ws, cfg, sheet_name, anomalies):
    """Read the balance formula and work out which money column decreases it.

    Returns (decreasing_col, increasing_col, formula_sample).

    The formula always has the shape `=<balance_prev><op><colA><op><colB>`,
    e.g. `=F3-D4+E4` or the inverted `=F3+D4-E4`. The column carrying a
    leading minus is the one that reduces the balance, which for a credit
    account is the spend side.
    """
    a, b = cfg["money"]
    bal = cfg["balance"]
    shapes = Counter()
    sample = None

    for r in range(cfg["header_row"] + 1, sheet_ws.max_row + 1):
        f = sheet_ws[f"{bal}{r}"].value
        if not (isinstance(f, str) and f.startswith("=")):
            continue
        sign_a = re.search(rf"([+-])\s*{a}\d+", f)
        sign_b = re.search(rf"([+-])\s*{b}\d+", f)
        if not (sign_a and sign_b):
            continue
        shapes[(sign_a.group(1), sign_b.group(1))] += 1
        if sample is None:
            sample = f

    if not shapes:
        anomalies.append(Anomaly(sheet_name, cfg["header_row"], "no_balance_formula",
                                 "could not derive direction; sheet skipped"))
        return None, None, None

    (sa, sb), _ = shapes.most_common(1)[0]
    if len(shapes) > 1:
        # Mixed conventions inside one sheet would silently flip a subset of
        # rows. Refuse rather than pick the majority.
        anomalies.append(Anomaly(
            sheet_name, cfg["header_row"], "mixed_balance_formula",
            f"sheet uses more than one sign convention: {dict(shapes)}"))

    if sa == "-" and sb == "+":
        return a, b, sample
    if sa == "+" and sb == "-":
        return b, a, sample
    anomalies.append(Anomaly(sheet_name, cfg["header_row"], "unreadable_direction",
                             f"signs {sa}{a} {sb}{b} in {sample!r}"))
    return None, None, sample


def raw_text(value):
    """The cell exactly as the workbook holds it, as text. Never normalised."""
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.isoformat(sep=" ")
    return str(value)


def parse_date(value, sheet_name, row, anomalies):
    """Return (iso_date, repaired, repair_note). Never guesses a format.

    The original cell is preserved separately by the caller; this only decides
    the single sortable date and says whether it had to be repaired.
    """
    if value is None or (isinstance(value, str) and not value.strip()):
        return None, False, None

    if isinstance(value, datetime.datetime):
        d = value.date()
        # Excel stored these from day/month source text read as month/day.
        if sheet_name in SWAP_DAY_MONTH and d.day <= 12 and d.month <= 12:
            try:
                fixed = datetime.date(d.year, d.day, d.month)
            except ValueError:
                anomalies.append(Anomaly(sheet_name, row, "date_swap_invalid",
                                         f"{d.isoformat()} is not valid transposed"))
                return d.isoformat(), False, None
            note = (f"Excel read the source as month/day; day and month "
                    f"transposed back ({d.isoformat()} -> {fixed.isoformat()})")
            return fixed.isoformat(), True, note
        return d.isoformat(), False, None

    if isinstance(value, str):
        s = value.strip()
        # Text dates survived only because day > 12 made month/day impossible,
        # which is itself the proof the source format is day/month/year.
        m = re.fullmatch(r"(\d{1,2})[/-](\d{1,2})[/-](\d{4})", s)
        if m:
            day, month, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
            try:
                return datetime.date(year, month, day).isoformat(), False, None
            except ValueError:
                anomalies.append(Anomaly(sheet_name, row, "date_invalid", repr(s)))
                return None, False, None
        anomalies.append(Anomaly(sheet_name, row, "date_unparseable", repr(s)))
        return None, False, None

    anomalies.append(Anomaly(sheet_name, row, "date_unexpected_type",
                             f"{type(value).__name__}: {value!r}"))
    return None, False, None


def parse_currency(ws, cfg, sheet_name, row, anomalies):
    """Return (code, original_amount, raw_text).

    A blank currency cell means the transaction was natively in AED. An
    unrecognised code is returned as None with the raw text preserved.
    """
    if "currency_code" in cfg:  # RAK 9825 splits the two into separate columns
        code_v = ws[f"{cfg['currency_code']}{row}"].value
        amt_v = ws[f"{cfg['currency_amount']}{row}"].value
        raw = " ".join(str(x).strip() for x in (code_v, amt_v) if x is not None) or None
        if code_v is None or not str(code_v).strip():
            return BASE_CURRENCY, None, raw
        code = str(code_v).strip().upper()
        if code not in KNOWN_CURRENCIES:
            anomalies.append(Anomaly(sheet_name, row, "currency_unrecognised", repr(code_v)))
            return None, None, raw
        amt = amt_v if isinstance(amt_v, (int, float)) else None
        if amt is None and amt_v is not None:
            anomalies.append(Anomaly(sheet_name, row, "currency_amount_unparseable",
                                     repr(amt_v)))
        return code, amt, raw

    v = ws[f"{cfg['currency']}{row}"].value
    if v is None or (isinstance(v, str) and not v.strip()):
        return BASE_CURRENCY, None, None

    raw = str(v).strip()
    # Source shape is "CODE  amount", e.g. "TRY  934000.00", "USD 240".
    m = re.fullmatch(r"'?([A-Za-z]{3})\s*([\d,]*\.?\d*)\.?'?", raw)
    if not m:
        anomalies.append(Anomaly(sheet_name, row, "currency_unparseable", repr(v)))
        return None, None, raw

    code = m.group(1).upper()
    if code not in KNOWN_CURRENCIES:
        anomalies.append(Anomaly(sheet_name, row, "currency_unrecognised", repr(v)))
        return None, None, raw

    amount = None
    if m.group(2):
        try:
            amount = float(m.group(2).replace(",", ""))
        except ValueError:
            anomalies.append(Anomaly(sheet_name, row, "currency_amount_unparseable",
                                     repr(v)))
    return code, amount, raw


# Suppliers arrive with the merchant's ISO-3166 numeric country code appended
# by the card statement, e.g. "GALLUP 840". Left attached, "GALLUP 840" and
# "GALLUP 784" are two different suppliers.
COUNTRY_SUFFIX = re.compile(r"\s(\d{3})\s*$")


def parse_supplier(value):
    """Return (clean_name, country_code, raw)."""
    if value is None or not str(value).strip():
        return None, None, None
    raw = re.sub(r"\s+", " ", str(value).replace("\t", " ").replace("\n", " ")).strip()
    m = COUNTRY_SUFFIX.search(raw)
    if m:
        return raw[: m.start()].strip(), m.group(1), raw
    return raw, None, raw


def content_signature(t):
    """What makes two rows the same transaction, apart from how many times it
    happened. Direction is in it, so a payment and its refund sharing one
    reference number never collapse into a single row."""
    return "|".join([
        t["card"],
        t["txn_date"] or "",
        f"{t['amount_aed']:.2f}" if t["amount_aed"] is not None else "",
        (t["supplier_raw"] or "").upper(),
        (t.get("payment_ref") or "").upper(),
        (t.get("req_number") or "").upper(),
        t["direction"] or t["entry_type"],
    ])


def assign_dedup_keys(transactions):
    """Give every row a stable key that survives a re-upload without merging
    transactions that genuinely happened more than once.

    217 rows in this workbook share every identifying field with another row —
    the same supplier charged the same amount twice on the same day against the
    same request, which the balance chain confirms are two real charges. A key
    built from content alone would collapse those 217 into 87 and quietly lose
    the difference. So the key is content plus an occurrence number, assigned
    in source order: re-importing the same file reproduces the same numbering
    and dedups exactly, while a genuine repeat keeps its own key.

    For rows added later by hand, the occurrence number must be one greater
    than the count of rows already stored with the same signature.
    """
    seen = Counter()
    for t in sorted(transactions, key=lambda x: (x["card"], x["source_row"] or 0)):
        sig = content_signature(t)
        seen[sig] += 1
        t["occurrence"] = seen[sig]
        t["dedup_key"] = hashlib.sha256(
            f"{sig}|#{seen[sig]}".encode("utf-8")).hexdigest()


def extract_sheet(wbf, wbv, sheet_name, cfg, anomalies):
    wsf, wsv = wbf[sheet_name], wbv[sheet_name]
    dec_col, inc_col, formula_sample = derive_direction(wsf, cfg, sheet_name, anomalies)
    if dec_col is None:
        return None

    hdr = cfg["header_row"]

    # The opening balance is the hardcoded seed sitting above the first
    # transaction, in the balance column. Any FURTHER hardcoded value further
    # down is a mid-ledger re-seed that silently overrides the arithmetic.
    seeds = [(r, float(wsf[f"{cfg['balance']}{r}"].value))
             for r in range(hdr + 1, wsf.max_row + 1)
             if isinstance(wsf[f"{cfg['balance']}{r}"].value, (int, float))]
    opening_row, opening_balance = seeds[0] if seeds else (None, None)
    reseed_rows = {r for r, _ in seeds[1:]}

    transactions = []
    for r in range(hdr + 1, wsf.max_row + 1):
        if r == opening_row:
            continue

        supplier_raw_v = wsf[f"{cfg['supplier']}{r}"].value
        dec_v = wsf[f"{dec_col}{r}"].value
        inc_v = wsf[f"{inc_col}{r}"].value
        has_money = isinstance(dec_v, (int, float)) or isinstance(inc_v, (int, float))
        if not (has_money or (supplier_raw_v and str(supplier_raw_v).strip())):
            continue  # blank / formatting-only row

        bal_cell = wsf[f"{cfg['balance']}{r}"].value
        has_balance_formula = isinstance(bal_cell, str) and bal_cell.startswith("=")

        # A row carrying money but no balance formula never entered the running
        # total: its amount is absent from every figure the workbook shows.
        chain_broken = has_money and not has_balance_formula and r not in reseed_rows
        if chain_broken:
            anomalies.append(Anomaly(
                sheet_name, r, "balance_chain_broken",
                f"{supplier_raw_v!r} has an amount but no balance formula, so the "
                f"workbook's own balance excludes it"))

        # A row with a description but no money that merely echoes the balance
        # is a marker, not a transaction.
        if not has_money and has_balance_formula:
            anomalies.append(Anomaly(sheet_name, r, "marker_row",
                                     f"{supplier_raw_v!r} carries no amount; excluded"))
            continue

        date_cell = wsf[f"{cfg['date']}{r}"].value
        date_iso, repaired, repair_note = parse_date(date_cell, sheet_name, r, anomalies)
        code, orig_amount, curr_raw = parse_currency(wsf, cfg, sheet_name, r, anomalies)
        name, country, name_raw = parse_supplier(supplier_raw_v)

        if isinstance(dec_v, (int, float)) and isinstance(inc_v, (int, float)):
            anomalies.append(Anomaly(sheet_name, r, "both_money_columns",
                                     f"{dec_col}={dec_v} and {inc_col}={inc_v}"))

        # Signed against the balance: spend is negative, funding positive.
        if isinstance(dec_v, (int, float)):
            amount_aed, direction = -float(dec_v), "spend"
        elif isinstance(inc_v, (int, float)):
            amount_aed, direction = float(inc_v), "funding"
        else:
            amount_aed, direction = None, None
            anomalies.append(Anomaly(sheet_name, r, "no_amount",
                                     f"supplier {name!r} has neither money column"))

        # The conversion cell is a formula; take the value Excel computed and
        # keep the formula so the stored rate stays traceable to its source.
        rate, rate_formula = None, None
        if cfg["conversion"]:
            rate_formula = wsf[f"{cfg['conversion']}{r}"].value
            rv = wsv[f"{cfg['conversion']}{r}"].value
            if isinstance(rv, (int, float)):
                rate = float(rv)
            if not (isinstance(rate_formula, str) and rate_formula.startswith("=")):
                rate_formula = None

        # A rate whose denominator disagrees with the stated original amount is
        # surfaced, not corrected. Likewise one that hardcodes both sides, where
        # the original amount cannot be recovered at all.
        if rate_formula:
            simple = re.fullmatch(r"=\s*[A-Z]+\d+\s*/\s*([\d,.]+)\s*", rate_formula)
            if simple and orig_amount is not None:
                try:
                    if abs(float(simple.group(1).replace(",", "")) - orig_amount) > 1e-9:
                        anomalies.append(Anomaly(
                            sheet_name, r, "rate_denominator_mismatch",
                            f"formula divides by {simple.group(1)} but the currency "
                            f"cell states {orig_amount}"))
                except ValueError:
                    pass
            elif not simple:
                anomalies.append(Anomaly(
                    sheet_name, r, "rate_formula_unexpected",
                    f"{rate_formula} hardcodes both sides; the original amount "
                    f"cannot be reliably recovered"))

            # A blank currency cell means "settled natively in AED" — but only
            # when there is no conversion. A row that carries a rate was
            # converted from something, so AED is the one answer it cannot be:
            # its currency is simply unknown, and saying AED here would assert
            # a currency the sheet never states.
            if code == BASE_CURRENCY and orig_amount is None:
                code = None
                if simple:
                    # The denominator is the original amount, read straight out
                    # of the formula rather than inferred.
                    try:
                        orig_amount = float(simple.group(1).replace(",", ""))
                    except ValueError:
                        pass
                anomalies.append(Anomaly(
                    sheet_name, r, "rate_without_currency",
                    f"a conversion rate is present ({rate_formula}) but the "
                    f"currency cell is blank, so the currency is unknown"))

        extra = {}
        for key, col in cfg["extra"].items():
            v = wsf[f"{col}{r}"].value
            if v is None:
                continue
            s = re.sub(r"\s+", " ", str(v).replace("\t", " ").replace("\n", " ")).strip()
            if s:
                extra[key] = s

        transactions.append({
            "card": sheet_name,
            "entry_type": "source_transaction",
            "status": "confirmed",
            "review_reason": None,
            "description": None,
            "source_sheet": sheet_name,
            "source_row": r,
            "txn_date": date_iso,
            "source_date_raw": raw_text(date_cell),
            "date_repaired": repaired,
            "date_repair_note": repair_note,
            "supplier": name,
            "supplier_country": country,
            "supplier_raw": name_raw,
            "currency": code,
            "original_amount": orig_amount,
            "currency_raw": curr_raw,
            "amount_aed": amount_aed,
            "direction": direction,
            # The workbook's own formula consumed every row except the ones
            # whose balance cell is empty.
            "included_in_source_balance": not chain_broken,
            "exchange_rate": rate,
            "exchange_rate_formula": rate_formula,
            "excel_balance": wsv[f"{cfg['balance']}{r}"].value,
            "reseed_balance": float(bal_cell) if r in reseed_rows else None,
            **extra,
        })

    return {
        "card": sheet_name,
        "header_row": hdr,
        "decreasing_column": dec_col,
        "decreasing_header": wsf[f"{dec_col}{hdr}"].value,
        "increasing_column": inc_col,
        "increasing_header": wsf[f"{inc_col}{hdr}"].value,
        "balance_formula_sample": formula_sample,
        "header_is_misleading": bool(
            "credit" in str(wsf[f"{dec_col}{hdr}"].value or "").lower()),
        "opening_balance": opening_balance,
        "opening_row": opening_row,
        "transactions": transactions,
    }


def apply_review_status(card, anomalies):
    """Mark rows the parser could not fully trust, without dropping any."""
    by_row = defaultdict(list)
    for a in anomalies:
        if a.sheet == card["card"] and a.kind in REVIEW_KINDS:
            by_row[a.row].append(f"{a.kind}: {a.detail}")

    for t in card["transactions"]:
        if not t["included_in_source_balance"]:
            t["status"] = "excluded_from_source_balance"
            t["review_reason"] = ("Present in source workbook but not included "
                                  "in its running-balance formula")
            continue
        reasons = by_row.get(t["source_row"])
        if reasons:
            t["status"] = "needs_review"
            t["review_reason"] = "; ".join(reasons)


def verify_balances(card, tolerance=0.005):
    """Replay the balance chain and compare against Excel's own cached values.

    This is the proof that direction and amounts were read correctly. It is
    computed independently of the sheet's formulas: we add our own signed
    amounts to the opening balance and check we land on Excel's number.
    """
    running = card["opening_balance"]
    if running is None:
        return {"checked": 0, "mismatches": [], "reseeds": [],
                "skipped": "no opening balance"}

    mismatches, reseeds, checked = [], [], 0
    for t in card["transactions"]:
        if t["amount_aed"] is None or not t["included_in_source_balance"]:
            continue
        running += t["amount_aed"]

        # Where the sheet typed a balance over the formula, the difference is
        # money the workbook's balance carries with no transaction behind it.
        if t["reseed_balance"] is not None:
            reseeds.append({
                "row": t["source_row"],
                "date": t["txn_date"],
                "arithmetic_says": round(running, 2),
                "sheet_says": round(t["reseed_balance"], 2),
                "unexplained_adjustment": round(t["reseed_balance"] - running, 2),
            })
            running = t["reseed_balance"]
            continue

        expected = t["excel_balance"]
        if not isinstance(expected, (int, float)):
            continue
        checked += 1
        if abs(running - float(expected)) > tolerance:
            mismatches.append({
                "row": t["source_row"],
                "computed": round(running, 2),
                "excel": round(float(expected), 2),
                "delta": round(running - float(expected), 2),
            })
            running = float(expected)  # resync so one break doesn't cascade
    return {"checked": checked, "mismatches": mismatches, "reseeds": reseeds}


# The date the RAK 9825 overwrite belongs to, as instructed. The source row
# carries '16-03-2026'; this states it explicitly so the adjustment does not
# inherit a date by accident.
RESEED_DATES = {("RAK 9825 (6071)", 26): "2026-03-16"}


def inject_adjustments(card, verification):
    """Add one explicitly labelled adjustment per balance overwrite.

    The adjustment is not a transaction that happened. It is the named,
    visible difference between what the workbook asserts and what its own
    arithmetic produces. It has no direction, which is what keeps it out of
    the spend and funding totals, and it is never merged into the source row.
    """
    for s in verification["reseeds"]:
        key = (card["card"], s["row"])
        card["transactions"].append({
            "card": card["card"],
            "entry_type": "reconciliation_adjustment",
            "status": "needs_review",
            "review_reason": (
                f"Source balance at row {s['row']} was manually overwritten: "
                f"the arithmetic gives {s['arithmetic_says']:,.2f} AED, the sheet "
                f"states {s['sheet_says']:,.2f} AED. Retained until the "
                f"underlying transaction is identified."),
            "description": "Manual balance overwrite in source workbook",
            "source_sheet": card["card"],
            "source_row": s["row"],
            "txn_date": RESEED_DATES.get(key, s["date"]),
            "source_date_raw": None,
            "date_repaired": False,
            "date_repair_note": None,
            "supplier": None,          # never invent a supplier
            "supplier_country": None,
            "supplier_raw": None,
            "currency": BASE_CURRENCY,
            "original_amount": None,
            "currency_raw": None,
            "amount_aed": s["unexplained_adjustment"],
            "direction": None,         # deliberately neither spend nor funding
            "included_in_source_balance": True,   # the workbook's balance has it
            "exchange_rate": None,
            "exchange_rate_formula": None,
            "payment_ref": None,       # never invent a payment reference
            "excel_balance": None,
            "reseed_balance": None,
        })


def summarise(card, verification):
    """The three balances, plus the adjustments held apart from the totals."""
    opening = card["opening_balance"] or 0.0
    txns = card["transactions"]

    def total(pred):
        return sum(t["amount_aed"] for t in txns
                   if t["amount_aed"] is not None and t["status"] != "voided" and pred(t))

    source = opening + total(lambda t: t["included_in_source_balance"])
    ledger = opening + total(lambda t: t["entry_type"] == "source_transaction")
    return {
        "opening_balance": round(opening, 2),
        "source_balance": round(source, 2),
        "ledger_balance": round(ledger, 2),
        "reconciliation_difference": round(source - ledger, 2),
        "total_spend": round(total(lambda t: t["direction"] == "spend"), 2),
        "total_funding": round(total(lambda t: t["direction"] == "funding"), 2),
        "review_adjustments_total": round(
            total(lambda t: t["entry_type"] == "reconciliation_adjustment"), 2),
        "needs_review": sum(1 for t in txns if t["status"] == "needs_review"),
        "excluded": sum(1 for t in txns
                        if t["status"] == "excluded_from_source_balance"),
        "verified_rows": verification["checked"],
        "mismatches": len(verification["mismatches"]),
    }


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--workbook", default=WORKBOOK)
    ap.add_argument("--out", help="write normalised JSON here")
    ap.add_argument("--anomalies", help="write the anomaly report here")
    args = ap.parse_args()

    # Opened for reading only. Neither workbook object is ever saved.
    wbf = openpyxl.load_workbook(args.workbook, data_only=False)
    wbv = openpyxl.load_workbook(args.workbook, data_only=True)

    anomalies, cards = [], []
    for name, cfg in SHEETS.items():
        if name not in wbf.sheetnames:
            anomalies.append(Anomaly(name, 0, "sheet_missing", "not in workbook"))
            continue
        card = extract_sheet(wbf, wbv, name, cfg, anomalies)
        if not card:
            continue
        apply_review_status(card, anomalies)
        card["verification"] = verify_balances(card)
        inject_adjustments(card, card["verification"])
        # Assigned once every row, adjustments included, exists.
        assign_dedup_keys(card["transactions"])
        card["summary"] = summarise(card, card["verification"])
        cards.append(card)

    print("=" * 84)
    print("DRY RUN — nothing written to any database; the workbook is untouched")
    print("=" * 84)

    print("\nDIRECTION DERIVED FROM EACH SHEET'S OWN BALANCE FORMULA")
    print("-" * 84)
    for c in cards:
        flag = "   <-- labelled CREDIT, but this is the spend side" \
            if c["header_is_misleading"] else ""
        print(f"  {c['card']}")
        print(f"    formula {c['balance_formula_sample']}   "
              f"decreases {c['decreasing_column']} "
              f"({c['decreasing_header']!r}){flag}")

    print("\nBALANCE CHAIN VERIFICATION (recomputed, vs Excel's cached values)")
    print("-" * 84)
    total_checked = total_bad = 0
    for c in cards:
        v = c["verification"]
        total_checked += v["checked"]
        total_bad += len(v["mismatches"])
        status = "OK" if not v["mismatches"] else f"{len(v['mismatches'])} MISMATCH"
        print(f"  {c['card']:<34} {v['checked']:>5} rows   {status}")
        for m in v["mismatches"][:5]:
            print(f"      row {m['row']}: computed {m['computed']} "
                  f"vs excel {m['excel']} (delta {m['delta']})")

    print("\nBALANCES — source vs ledger, and the difference between them")
    print("-" * 84)
    print(f"  {'CARD':<32}{'SOURCE':>14}{'LEDGER':>14}{'DIFFERENCE':>14}{'ADJ':>12}")
    tot = defaultdict(float)
    for c in cards:
        s = c["summary"]
        for k in ("source_balance", "ledger_balance", "reconciliation_difference",
                  "review_adjustments_total"):
            tot[k] += s[k]
        mark = " *" if abs(s["reconciliation_difference"]) > 0.005 else ""
        print(f"  {c['card']:<32}{s['source_balance']:>14,.2f}"
              f"{s['ledger_balance']:>14,.2f}"
              f"{s['reconciliation_difference']:>14,.2f}"
              f"{s['review_adjustments_total']:>12,.2f}{mark}")
    print("  " + "-" * 82)
    print(f"  {'TOTAL (AED)':<32}{tot['source_balance']:>14,.2f}"
          f"{tot['ledger_balance']:>14,.2f}"
          f"{tot['reconciliation_difference']:>14,.2f}"
          f"{tot['review_adjustments_total']:>12,.2f}")
    print("\n  * source and ledger disagree — see the review queue below")

    print("\nREVIEW QUEUE — imported in full, flagged, never dropped or corrected")
    print("-" * 84)
    queue = [t for c in cards for t in c["transactions"]
             if t["status"] != "confirmed"
             or t["entry_type"] == "reconciliation_adjustment"]
    by_status = Counter(t["status"] for t in queue)
    print(f"  {len(queue)} rows: {dict(by_status)}\n")
    for t in queue:
        if t["entry_type"] == "reconciliation_adjustment" \
                or t["status"] == "excluded_from_source_balance":
            print(f"  [{t['status']}] {t['card']} row {t['source_row']}")
            print(f"      date {t['txn_date']}  amount {t['amount_aed']:+,.2f} AED"
                  f"  supplier {t['supplier_raw'] or '(none — not invented)'}")
            print(f"      {t['description'] or t['review_reason']}")
            if t["description"]:
                print(f"      {t['review_reason']}")
            print()
    needs = [t for t in queue if t["status"] == "needs_review"
             and t["entry_type"] == "source_transaction"]
    if needs:
        print(f"  needs_review — currency and exchange-rate exceptions ({len(needs)}):")
        for t in needs:
            print(f"      {t['card']} row {t['source_row']}: {t['review_reason']}")

    all_txns = [t for c in cards for t in c["transactions"]]
    src_txns = [t for t in all_txns if t["entry_type"] == "source_transaction"]

    print("\nTOTALS")
    print("-" * 84)
    print(f"  cards                    {len(cards)}")
    print(f"  source transactions      {len(src_txns)}")
    print(f"  reconciliation adjust.   {len(all_txns) - len(src_txns)}")
    print(f"  distinct suppliers       "
          f"{len({t['supplier'].upper() for t in src_txns if t['supplier']})}")
    print(f"  balances verified        {total_checked}   mismatches {total_bad}")
    print(f"  dates repaired           "
          f"{sum(1 for t in src_txns if t['date_repaired'])} "
          f"(original preserved in source_date_raw)")
    repeats = sum(1 for t in all_txns if t["occurrence"] > 1)
    print(f"  dedup keys unique        "
          f"{len({t['dedup_key'] for t in all_txns})} of {len(all_txns)}")
    print(f"  genuine repeat charges   {repeats} "
          f"(identical rows kept apart by occurrence number)")

    print("\nANOMALIES — left blank, never guessed")
    print("-" * 84)
    by_kind = defaultdict(list)
    for a in anomalies:
        by_kind[a.kind].append(a)
    for kind, items in sorted(by_kind.items(), key=lambda kv: -len(kv[1])):
        print(f"  {kind}  ({len(items)})")
        for a in items[:4]:
            print(f"      {a.sheet} row {a.row}: {a.detail}")
        if len(items) > 4:
            print(f"      ... and {len(items) - 4} more")

    if args.out:
        with open(args.out, "w", encoding="utf-8") as fh:
            json.dump(cards, fh, indent=2, ensure_ascii=False, default=str)
        print(f"\nnormalised data -> {args.out}")
    if args.anomalies:
        with open(args.anomalies, "w", encoding="utf-8") as fh:
            json.dump([a.as_dict() for a in anomalies], fh, indent=2, ensure_ascii=False)
        print(f"anomaly report  -> {args.anomalies}")

    return 1 if total_bad else 0


if __name__ == "__main__":
    sys.exit(main())

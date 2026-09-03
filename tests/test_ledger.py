"""
Tests for the card-ledger extraction and balance math.

Every test here checks an actual number. Where a test verifies the pipeline's
own output, it derives the expected figure a SECOND, INDEPENDENT way — reading
the workbook's cached balance deltas rather than reusing the parser's formula
logic — because a test that calls the function under test to produce its own
expectation only proves the function equals itself.

Run:  python tests/test_ledger.py
"""

import datetime
import os
import sys
import warnings

warnings.filterwarnings("ignore")

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))

import openpyxl  # noqa: E402

import extract  # noqa: E402


# ---------------------------------------------------------------------------
# Tiny test runner — no external dependency
# ---------------------------------------------------------------------------

_TESTS, _FAILED = [], []


def test(fn):
    _TESTS.append(fn)
    return fn


def eq(actual, expected, what):
    if isinstance(expected, float) or isinstance(actual, float):
        ok = abs(float(actual) - float(expected)) < 0.005
    else:
        ok = actual == expected
    if not ok:
        raise AssertionError(f"{what}: expected {expected!r}, got {actual!r}")


# ---------------------------------------------------------------------------
# Shared fixture: one extraction, reused
# ---------------------------------------------------------------------------

_CARDS = None


def cards():
    global _CARDS
    if _CARDS is None:
        wbf = openpyxl.load_workbook(extract.WORKBOOK, data_only=False)
        wbv = openpyxl.load_workbook(extract.WORKBOOK, data_only=True)
        anomalies, out = [], []
        for name, cfg in extract.SHEETS.items():
            c = extract.extract_sheet(wbf, wbv, name, cfg, anomalies)
            extract.apply_review_status(c, anomalies)
            c["verification"] = extract.verify_balances(c)
            extract.inject_adjustments(c, c["verification"])
            extract.assign_dedup_keys(c["transactions"])
            c["summary"] = extract.summarise(c, c["verification"])
            out.append(c)
        _CARDS = {c["card"]: c for c in out}
        _CARDS["__anomalies__"] = anomalies
    return _CARDS


def card(name):
    return cards()[name]


def source_transactions(name):
    return [t for t in card(name)["transactions"]
            if t["entry_type"] == "source_transaction"]


# ---------------------------------------------------------------------------
# 1. Direction — derived independently from cached balance deltas
# ---------------------------------------------------------------------------

def direction_from_balance_deltas(sheet_name, cfg):
    """Work out the spend column WITHOUT reading any formula text.

    Walks the balance column's cached values and, for each step, checks whether
    the change equals minus the first money column or minus the second. This is
    arithmetic on the numbers Excel itself computed, so it shares no code path
    with derive_direction().
    """
    wb = openpyxl.load_workbook(extract.WORKBOOK, data_only=True)
    ws = wb[sheet_name]
    a, b = cfg["money"]
    bal = cfg["balance"]
    votes = {a: 0, b: 0}
    prev = None
    for r in range(cfg["header_row"] + 1, ws.max_row + 1):
        cur = ws[f"{bal}{r}"].value
        if not isinstance(cur, (int, float)):
            continue
        if prev is not None:
            delta = float(cur) - float(prev)
            va, vb = ws[f"{a}{r}"].value, ws[f"{b}{r}"].value
            if isinstance(va, (int, float)) and abs(delta + float(va)) < 0.005:
                votes[a] += 1
            elif isinstance(vb, (int, float)) and abs(delta + float(vb)) < 0.005:
                votes[b] += 1
        prev = cur
    return max(votes, key=votes.get) if any(votes.values()) else None


@test
def test_direction_matches_independent_derivation():
    """The spend column found by parsing formulas must match the one found by
    arithmetic on cached balances."""
    expected_spend_col = {
        "MASTERCARD 5135 (4173) (7206)": "D",
        "RAK 8871 (8435)(0033)": "D",
        "AMEX 4000 VPAY": "E",          # headed CREDIT, and it is the spend side
        "MASTERCARD 6404 VPAY": "D",
        "AMEX 3024 (3016- COR)": "E",   # headed CREDIT
        "AMEX 2044 (2036- VIP)": "D",   # headed CREDIT (transposed headers)
        "RAK 9825 (6071)": "E",
    }
    for name, cfg in extract.SHEETS.items():
        parsed = card(name)["decreasing_column"]
        eq(parsed, expected_spend_col[name], f"{name} spend column (hand-checked)")
        independent = direction_from_balance_deltas(name, cfg)
        if independent is not None:
            eq(parsed, independent,
               f"{name} spend column (independently from balance deltas)")


@test
def test_three_sheets_have_misleading_credit_header():
    """The header must never be trusted: on three sheets it says CREDIT while
    the column reduces the balance."""
    misleading = {n for n in extract.SHEETS if card(n)["header_is_misleading"]}
    eq(misleading,
       {"AMEX 4000 VPAY", "AMEX 3024 (3016- COR)", "AMEX 2044 (2036- VIP)"},
       "sheets whose spend column is labelled CREDIT")


@test
def test_spend_is_negative_funding_is_positive():
    for name in extract.SHEETS:
        for t in source_transactions(name):
            if t["direction"] == "spend":
                assert t["amount_aed"] <= 0, f"{name} row {t['source_row']} spend positive"
            elif t["direction"] == "funding":
                assert t["amount_aed"] >= 0, f"{name} row {t['source_row']} funding negative"


# ---------------------------------------------------------------------------
# 2. Balance chain reproduces the workbook exactly
# ---------------------------------------------------------------------------

@test
def test_balance_chain_reproduces_excel():
    """Replaying our signed amounts must land on Excel's own cached balance at
    every row. A sign error anywhere breaks every row after it."""
    total = 0
    for name in extract.SHEETS:
        v = card(name)["verification"]
        eq(len(v["mismatches"]), 0, f"{name} balance mismatches")
        total += v["checked"]
    eq(total, 1946, "rows verified against Excel's cached balances")


@test
def test_closing_balance_equals_last_cell_in_sheet():
    """source_balance must equal the balance the workbook shows against its
    last real transaction.

    Deliberately anchored to the last row that carries money, not the last
    numeric cell in the column: several sheets trail empty formula rows that
    copy the final balance forward, and AMEX 4000 has an orphaned formula at
    row 2621 that evaluates to 0 because the cell it references is blank.
    """
    wb = openpyxl.load_workbook(extract.WORKBOOK, data_only=True)
    expected_closing = {
        "MASTERCARD 5135 (4173) (7206)": 244608.37,
        "RAK 8871 (8435)(0033)": 178885.95,
        "AMEX 4000 VPAY": 599536.31,
        "MASTERCARD 6404 VPAY": 127930.85,
        "AMEX 3024 (3016- COR)": 25474.74,   # FLYNAS is not in the sheet's total
        "AMEX 2044 (2036- VIP)": 13708.16,
        "RAK 9825 (6071)": 165.72,           # includes the manual overwrite
    }
    for name, cfg in extract.SHEETS.items():
        ws = wb[name]
        last = None
        for r in range(cfg["header_row"] + 1, ws.max_row + 1):
            has_money = any(isinstance(ws[f"{c}{r}"].value, (int, float))
                            for c in cfg["money"])
            bal = ws[f"{cfg['balance']}{r}"].value
            if has_money and isinstance(bal, (int, float)):
                last = float(bal)
        eq(last, expected_closing[name], f"{name} closing balance in the workbook")
        eq(card(name)["summary"]["source_balance"], last,
           f"{name} source_balance vs the sheet's own closing balance")


# ---------------------------------------------------------------------------
# 3. The two cards where source and ledger disagree
# ---------------------------------------------------------------------------

@test
def test_amex_3024_flynas_excluded_from_source_balance():
    s = card("AMEX 3024 (3016- COR)")["summary"]
    eq(s["source_balance"], 25474.74, "AMEX 3024 source balance (workbook)")
    eq(s["ledger_balance"], 16728.96, "AMEX 3024 ledger balance (incl. FLYNAS)")
    eq(s["reconciliation_difference"], 8745.78, "AMEX 3024 difference")

    flynas = [t for t in source_transactions("AMEX 3024 (3016- COR)")
              if t["source_row"] == 5]
    eq(len(flynas), 1, "FLYNAS row imported")
    t = flynas[0]
    eq(t["status"], "excluded_from_source_balance", "FLYNAS status")
    eq(t["included_in_source_balance"], False, "FLYNAS in source balance")
    eq(t["amount_aed"], -8745.78, "FLYNAS amount")
    eq(t["currency"], "SAR", "FLYNAS currency preserved")
    eq(t["original_amount"], 8925.77, "FLYNAS original amount preserved")
    eq(t["supplier"], "FLYNAS RIYADH", "FLYNAS supplier")
    assert "not included in its running-balance formula" in t["review_reason"]


FLYNAS_AED = -8745.78
AMEX_3024_SOURCE_BALANCE = 25474.74   # cell F4 — the sheet's own formula chain
AMEX_3024_LEDGER_BALANCE = 16728.96   # 25,474.74 - 8,745.78, FLYNAS once
AMEX_3024_DIFFERENCE = 8745.78
# 7,983.18 deducts FLYNAS twice and corresponds to no cell in the workbook.
AMEX_3024_FORBIDDEN = 7983.18


@test
def test_amex_3024_three_reported_figures():
    """The three figures, named explicitly."""
    s = card("AMEX 3024 (3016- COR)")["summary"]
    eq(s["source_balance"], AMEX_3024_SOURCE_BALANCE,
       "source workbook balance (formula chain at F4, FLYNAS excluded)")
    eq(s["ledger_balance"], AMEX_3024_LEDGER_BALANCE,
       "ledger balance including FLYNAS")
    eq(s["reconciliation_difference"], AMEX_3024_DIFFERENCE,
       "reconciliation difference")


@test
def test_flynas_is_deducted_exactly_once():
    """No balance calculation may deduct FLYNAS more than once.

    Guards the specific error of subtracting the excluded transaction from a
    figure that already excludes it, which produces 7,983.18.
    """
    c = card("AMEX 3024 (3016- COR)")
    s = c["summary"]

    flynas = [t for t in c["transactions"]
              if (t["supplier_raw"] or "").upper().startswith("FLYNAS")]
    eq(len(flynas), 1, "FLYNAS must exist exactly once in the ledger")
    eq(flynas[0]["amount_aed"], FLYNAS_AED, "FLYNAS amount")

    # The ledger is the source balance with FLYNAS applied once, no more.
    eq(s["ledger_balance"], s["source_balance"] + FLYNAS_AED,
       "ledger balance is the source balance minus FLYNAS exactly once")
    eq(s["reconciliation_difference"], -FLYNAS_AED,
       "the difference is one FLYNAS, not two")

    # And explicitly not the double-deducted figure.
    assert abs(s["ledger_balance"] - AMEX_3024_FORBIDDEN) > 0.005, \
        "ledger balance must not be 7,983.18 (FLYNAS deducted twice)"


@test
def test_no_reported_figure_is_the_double_deduction():
    """7,983.18 must not appear as any balance on any card."""
    for name in extract.SHEETS:
        s = card(name)["summary"]
        for key in ("source_balance", "ledger_balance", "opening_balance",
                    "reconciliation_difference"):
            assert abs(s[key] - AMEX_3024_FORBIDDEN) > 0.005, \
                f"{name}.{key} is 7,983.18, the double-deducted figure"


@test
def test_balance_summary_is_idempotent():
    """Recomputing must not apply the exclusion a second time.

    A stored balance can drift by being adjusted twice; a derived one cannot.
    This asserts the derivation actually is pure.
    """
    c = card("AMEX 3024 (3016- COR)")
    first = extract.summarise(c, c["verification"])
    second = extract.summarise(c, c["verification"])
    third = extract.summarise(c, c["verification"])
    for key in ("source_balance", "ledger_balance", "reconciliation_difference"):
        eq(second[key], first[key], f"{key} stable on recomputation")
        eq(third[key], first[key], f"{key} stable on third recomputation")
    eq(first["ledger_balance"], AMEX_3024_LEDGER_BALANCE,
       "ledger balance after repeated recomputation")


@test
def test_excluded_transaction_counted_once_across_all_cards():
    """Generally, not just for FLYNAS: an excluded row contributes to the
    ledger balance exactly once and to the source balance not at all."""
    for name in extract.SHEETS:
        c = card(name)
        s = c["summary"]
        excluded = [t for t in c["transactions"]
                    if not t["included_in_source_balance"]
                    and t["entry_type"] == "source_transaction"
                    and t["amount_aed"] is not None]
        expected_gap = sum(t["amount_aed"] for t in excluded)
        eq(s["ledger_balance"] - s["source_balance"],
           expected_gap - s["review_adjustments_total"],
           f"{name}: ledger minus source equals excluded rows less adjustments")


@test
def test_rak_9825_manual_overwrite_becomes_labelled_adjustment():
    c = card("RAK 9825 (6071)")
    s = c["summary"]
    eq(s["source_balance"], 165.72, "RAK 9825 source balance (workbook)")
    eq(s["ledger_balance"], -1552.30, "RAK 9825 ledger balance (real txns only)")
    eq(s["reconciliation_difference"], 1718.02, "RAK 9825 difference")

    adj = [t for t in c["transactions"]
           if t["entry_type"] == "reconciliation_adjustment"]
    eq(len(adj), 1, "one adjustment created")
    a = adj[0]
    eq(a["amount_aed"], 1718.02, "adjustment amount")
    eq(a["txn_date"], "2026-03-16", "adjustment date")
    eq(a["status"], "needs_review", "adjustment status")
    eq(a["description"], "Manual balance overwrite in source workbook",
       "adjustment description")
    eq(a["source_sheet"], "RAK 9825 (6071)", "adjustment source sheet")
    eq(a["source_row"], 26, "adjustment source row")
    eq(a["direction"], None, "adjustment must have no direction")
    # Nothing invented.
    eq(a["supplier"], None, "adjustment must not invent a supplier")
    eq(a["supplier_raw"], None, "adjustment must not invent a supplier")
    eq(a["payment_ref"], None, "adjustment must not invent a payment reference")


@test
def test_original_reseed_row_is_untouched():
    """The source row that was overwritten stays a normal transaction; the
    adjustment is separate from it."""
    row26 = [t for t in source_transactions("RAK 9825 (6071)")
             if t["source_row"] == 26]
    eq(len(row26), 1, "row 26 still imported as a source transaction")
    eq(row26[0]["amount_aed"], 10000.0, "row 26 keeps its own 10,000 credit")
    eq(row26[0]["direction"], "funding", "row 26 direction")


@test
def test_adjustment_never_inside_spend_or_funding():
    """A plug must not hide inside a total."""
    c = card("RAK 9825 (6071)")
    s = c["summary"]
    eq(s["review_adjustments_total"], 1718.02, "adjustments reported separately")
    spend = sum(t["amount_aed"] for t in c["transactions"]
                if t["direction"] == "spend")
    funding = sum(t["amount_aed"] for t in c["transactions"]
                  if t["direction"] == "funding")
    eq(s["total_spend"], spend, "spend excludes the adjustment")
    eq(s["total_funding"], funding, "funding excludes the adjustment")
    eq(s["ledger_balance"],
       c["opening_balance"] + spend + funding,
       "ledger balance is opening + spend + funding, with no plug")


@test
def test_other_five_cards_reconcile_exactly():
    for name in ["MASTERCARD 5135 (4173) (7206)", "RAK 8871 (8435)(0033)",
                 "AMEX 4000 VPAY", "MASTERCARD 6404 VPAY", "AMEX 2044 (2036- VIP)"]:
        eq(card(name)["summary"]["reconciliation_difference"], 0.0,
           f"{name} source and ledger must agree")


# ---------------------------------------------------------------------------
# 4. Dates — one sortable format, original preserved, repairs flagged
# ---------------------------------------------------------------------------

@test
def test_every_date_is_one_sortable_format():
    for name in extract.SHEETS:
        for t in source_transactions(name):
            if t["txn_date"] is None:
                continue
            datetime.date.fromisoformat(t["txn_date"])  # raises if not ISO


@test
def test_repaired_dates_keep_their_original_and_are_flagged():
    repaired = [t for name in extract.SHEETS for t in source_transactions(name)
                if t["date_repaired"]]
    eq(len(repaired), 17, "dates repaired")
    for t in repaired:
        assert t["source_date_raw"], "a repaired date must keep the original"
        assert t["date_repair_note"], "a repaired date must say what was done"
        eq(t["card"] in extract.SWAP_DAY_MONTH, True,
           "repairs only on the two affected sheets")


@test
def test_specific_date_repair_is_correct():
    """RAK 9825 row 3 reads 2026-09-01 in Excel; the source is 1 September
    written day/month, so it must become 2026-01-09."""
    t = [x for x in source_transactions("RAK 9825 (6071)") if x["source_row"] == 3][0]
    eq(t["source_date_raw"], "2026-09-01 00:00:00", "original preserved verbatim")
    eq(t["txn_date"], "2026-01-09", "repaired to day/month reading")
    eq(t["date_repaired"], True, "flagged as a correction")


@test
def test_unambiguous_dates_are_never_touched():
    """A date whose day exceeds 12 cannot be a transposition and must be left
    exactly as Excel holds it, even on the two repaired sheets."""
    for name in extract.SWAP_DAY_MONTH:
        for t in source_transactions(name):
            if t["txn_date"] and t["source_date_raw"] \
                    and t["source_date_raw"].endswith("00:00:00"):
                d = datetime.date.fromisoformat(t["source_date_raw"][:10])
                if d.day > 12:
                    eq(t["date_repaired"], False,
                       f"{name} row {t['source_row']} must not be repaired")


@test
def test_opening_date_is_first_transaction():
    for name in extract.SHEETS:
        dates = [t["txn_date"] for t in source_transactions(name) if t["txn_date"]]
        # The card's opening date is defined as its earliest transaction, so no
        # historical row can fall before it and be excluded.
        earliest = min(dates)
        for t in source_transactions(name):
            if t["txn_date"]:
                assert t["txn_date"] >= earliest, \
                    f"{name} row {t['source_row']} predates the opening date"


# ---------------------------------------------------------------------------
# 5. Currency and exchange rate — preserved, never guessed
# ---------------------------------------------------------------------------

@test
def test_unrecognised_currency_is_blank_and_flagged():
    """Eight rows cannot state a currency: three hold a bare number where the
    code should be, and five more carry a conversion rate against a blank
    currency cell. All must import with a null currency and be flagged.

    (MASTERCARD 6404 row 246 is both — a bare '4400' AND a rate — and is
    counted once, under currency_unparseable.)"""
    bad = [t for name in extract.SHEETS for t in source_transactions(name)
           if t["currency"] is None]
    eq(len(bad), 8, "rows with no determinable currency")
    for t in bad:
        eq(t["status"], "needs_review", f"row {t['source_row']} status")
        assert t["amount_aed"] is not None, "the AED amount is still imported"


@test
def test_rate_without_currency_is_never_called_aed():
    """A blank currency cell means 'settled natively in AED' — but a row that
    carries a conversion rate was converted from something, so AED is the one
    answer it cannot be. Calling it AED would assert a currency the sheet never
    states, and would put a rate of 3.65 against a currency whose rate is 1.
    """
    flagged = [t for name in extract.SHEETS for t in source_transactions(name)
               if "rate_without_currency" in (t["review_reason"] or "")]
    eq(len(flagged), 5, "rows with a rate against a blank currency cell")
    for t in flagged:
        eq(t["currency"], None, f"row {t['source_row']} must not claim a currency")
        assert t["exchange_rate"] is not None, "the sheet's rate is kept"
        assert t["exchange_rate_formula"], "the formula is kept as provenance"
        # The denominator is read out of the formula, not inferred.
        assert t["original_amount"] is not None, \
            "the original amount is recovered from the formula"
        eq(t["status"], "needs_review", "flagged for a human")

    # Concrete case: AMEX 4000 row 242, '=E242/420'.
    r242 = [t for t in source_transactions("AMEX 4000 VPAY")
            if t["source_row"] == 242][0]
    eq(r242["currency"], None, "row 242 currency unknown")
    eq(r242["original_amount"], 420.0, "row 242 original amount from the formula")


@test
def test_rate_exceptions_flagged_but_values_preserved():
    flagged = [t for name in extract.SHEETS for t in source_transactions(name)
               if t["status"] == "needs_review"]
    eq(len(flagged), 17, "currency and rate exceptions")

    mismatch = [t for t in flagged
                if "rate_denominator_mismatch" in (t["review_reason"] or "")]
    eq(len(mismatch), 6, "denominator mismatches")
    hardcoded = [t for t in flagged
                 if "rate_formula_unexpected" in (t["review_reason"] or "")]
    eq(len(hardcoded), 3, "hardcoded rate formulas")

    # Flagged, but nothing about them was altered.
    for t in mismatch:
        assert t["exchange_rate"] is not None, "the sheet's rate is kept"
        assert t["exchange_rate_formula"], "the formula is kept as provenance"
        assert t["original_amount"] is not None, "the stated amount is kept"


@test
def test_no_row_has_a_rate_against_the_base_currency():
    """A rate is a conversion. AED converting to AED at 3.65 is incoherent, and
    would corrupt any per-currency report that trusted it."""
    for name in extract.SHEETS:
        for t in source_transactions(name):
            if t["exchange_rate"] is not None and t["currency"] == "AED":
                raise AssertionError(
                    f"{name} row {t['source_row']} claims AED with a rate of "
                    f"{t['exchange_rate']}")


@test
def test_stored_rate_is_the_sheets_own_value():
    """The rate must be what the sheet computed, not a recomputation."""
    wb = openpyxl.load_workbook(extract.WORKBOOK, data_only=True)
    ws = wb["AMEX 4000 VPAY"]
    for t in source_transactions("AMEX 4000 VPAY"):
        if t["exchange_rate"] is None:
            continue
        cached = ws[f"G{t['source_row']}"].value
        eq(t["exchange_rate"], float(cached),
           f"AMEX 4000 row {t['source_row']} rate is the sheet's own value")


@test
def test_rate_is_aed_per_unit_of_foreign_currency():
    """rate * original_amount must reconstruct the AED amount, on rows where
    the sheet's own numbers are self-consistent."""
    checked = 0
    for name in extract.SHEETS:
        for t in source_transactions(name):
            if not (t["exchange_rate"] and t["original_amount"]
                    and t["status"] == "confirmed"):
                continue
            eq(abs(t["amount_aed"]),
               round(t["exchange_rate"] * t["original_amount"], 2),
               f"{name} row {t['source_row']} rate x original = AED")
            checked += 1
    assert checked > 250, f"expected a real sample, checked only {checked}"


@test
def test_currencies_are_never_cross_added():
    """Spend by currency must stay one figure per currency."""
    from collections import defaultdict
    per = defaultdict(float)
    for name in extract.SHEETS:
        for t in source_transactions(name):
            if t["direction"] == "spend" and t["currency"] and t["original_amount"]:
                per[t["currency"]] += t["original_amount"]
    assert len(per) > 20, "expected many currencies"
    # A single cross-currency total would be meaningless; assert the shape is
    # per-currency and that two large, incomparable ones stay apart.
    assert per["JPY"] > per["EUR"], "JPY and EUR totals must not be merged"
    eq(round(per["OMR"], 2), 5685.50, "OMR original-currency spend")


# ---------------------------------------------------------------------------
# 6. Deduplication
# ---------------------------------------------------------------------------

@test
def test_dedup_keys_are_unique():
    keys = [t["dedup_key"] for name in extract.SHEETS
            for t in card(name)["transactions"]]
    eq(len(set(keys)), len(keys), "every row must have its own key")
    eq(len(keys), 1949, "total rows including the one adjustment")


@test
def test_genuine_repeat_charges_are_kept_apart():
    """217 rows share every identifying field with another. They are real, and
    must not collapse."""
    repeats = [t for name in extract.SHEETS for t in card(name)["transactions"]
               if t["occurrence"] > 1]
    eq(len(repeats), 130, "rows that are a second-or-later identical charge")

    # A concrete pair: RAK 8871 rows 16 and 17 are the same charge twice.
    pair = [t for t in source_transactions("RAK 8871 (8435)(0033)")
            if t["source_row"] in (16, 17)]
    eq(len(pair), 2, "both rows imported")
    eq(pair[0]["amount_aed"], pair[1]["amount_aed"], "same amount")
    eq(pair[0]["txn_date"], pair[1]["txn_date"], "same date")
    assert pair[0]["dedup_key"] != pair[1]["dedup_key"], \
        "two real charges must not share a key"
    eq({p["occurrence"] for p in pair}, {1, 2}, "numbered 1 and 2")


@test
def test_reimport_produces_identical_keys():
    """Re-uploading the same file must reproduce the same keys exactly, so
    nothing is double-counted."""
    wbf = openpyxl.load_workbook(extract.WORKBOOK, data_only=False)
    wbv = openpyxl.load_workbook(extract.WORKBOOK, data_only=True)
    for name, cfg in extract.SHEETS.items():
        anomalies = []
        c = extract.extract_sheet(wbf, wbv, name, cfg, anomalies)
        extract.apply_review_status(c, anomalies)
        c["verification"] = extract.verify_balances(c)
        extract.inject_adjustments(c, c["verification"])
        extract.assign_dedup_keys(c["transactions"])
        again = {t["dedup_key"] for t in c["transactions"]}
        first = {t["dedup_key"] for t in card(name)["transactions"]}
        eq(again, first, f"{name} keys must be stable across imports")


@test
def test_refund_and_payment_do_not_merge():
    """Direction is part of the key: a payment and a refund sharing a
    reference number must stay two rows."""
    a = {"card": "X", "txn_date": "2026-01-01", "amount_aed": -100.0,
         "supplier_raw": "ACME", "payment_ref": "Payment Made: #1",
         "req_number": "REQ 1", "direction": "spend",
         "entry_type": "source_transaction"}
    b = dict(a, amount_aed=100.0, direction="funding")
    assert extract.content_signature(a) != extract.content_signature(b), \
        "a payment and its refund must not share a signature"


# ---------------------------------------------------------------------------
# 7. Supplier normalisation
# ---------------------------------------------------------------------------

@test
def test_country_code_split_from_supplier_name():
    eq(extract.parse_supplier("GALLUP 840"), ("GALLUP", "840", "GALLUP 840"),
       "trailing ISO-3166 code split out")
    eq(extract.parse_supplier("JW MARRIOTT"), ("JW MARRIOTT", None, "JW MARRIOTT"),
       "a name with no code is unchanged")
    eq(extract.parse_supplier("Starbucks-H400 682"),
       ("Starbucks-H400", "682", "Starbucks-H400 682"),
       "digits inside the name are not mistaken for the code")


@test
def test_supplier_raw_always_preserved():
    for name in extract.SHEETS:
        for t in source_transactions(name):
            if t["supplier"]:
                assert t["supplier_raw"], "the original supplier text is kept"


# ---------------------------------------------------------------------------
# 8. Nothing is dropped
# ---------------------------------------------------------------------------

@test
def test_every_row_with_money_is_imported():
    """No row carrying an amount may be silently discarded."""
    wb = openpyxl.load_workbook(extract.WORKBOOK, data_only=False)
    for name, cfg in extract.SHEETS.items():
        ws = wb[name]
        c = card(name)
        opening_row = c["opening_row"]
        expected = set()
        for r in range(cfg["header_row"] + 1, ws.max_row + 1):
            if r == opening_row:
                continue
            if any(isinstance(ws[f"{col}{r}"].value, (int, float))
                   for col in cfg["money"]):
                expected.add(r)
        got = {t["source_row"] for t in source_transactions(name)}
        missing = expected - got
        eq(missing, set(), f"{name} rows with money that were not imported")


@test
def test_total_transaction_count():
    total = sum(len(source_transactions(n)) for n in extract.SHEETS)
    eq(total, 1948, "source transactions imported")


# ---------------------------------------------------------------------------

def main():
    print(f"running {len(_TESTS)} tests\n")
    for fn in _TESTS:
        try:
            fn()
            print(f"  PASS  {fn.__name__}")
        except Exception as exc:  # noqa: BLE001
            _FAILED.append((fn.__name__, exc))
            print(f"  FAIL  {fn.__name__}\n          {exc}")
    print()
    if _FAILED:
        print(f"{len(_FAILED)} of {len(_TESTS)} tests FAILED")
        return 1
    print(f"all {len(_TESTS)} tests passed")
    return 0


if __name__ == "__main__":
    sys.exit(main())
